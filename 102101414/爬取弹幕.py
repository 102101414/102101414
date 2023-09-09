import requests
import re
from lxml import etree
import heapq
import openpyxl
cnts={}
for page in range(1,11):
    # 爬取一页视频的url
    url=''
    if page==1:
        url=f'https://search.bilibili.com/all?keyword=日本核污染水排海'
    else:
        url=f'https://search.bilibili.com/all?vt=68973445&keyword=日本核污染水排海&page={page}'
    print(page)
    headers = {
    'cookie':'nostalgia_conf=-1; _uuid=DBA108CF10-7B18-31D2-F235-6A7D287698B813584infoc; buvid3=E13C5F69-3CDE-3643-28D3-53551C09952A14835infoc; b_nut=1666596615; buvid4=EF91F08C-3200-880E-183C-32A8153B8D3314835-022102415-V35mpzdvTWSkhBz1UroM1g%3D%3D; i-wanna-go-back=-1; b_ut=7; is-2022-channel=1; FEED_LIVE_VERSION=V8; header_theme_version=CLOSE; home_feed_column=5; CURRENT_BLACKGAP=0; rpdid=|(u))kkYuuu|0J\'uY))J~kuum; CURRENT_FNVAL=4048; browser_resolution=1494-789; fingerprint=722658bd40bab4c1be58d7e8b4e4fcfe; buvid_fp_plain=undefined; buvid_fp=722658bd40bab4c1be58d7e8b4e4fcfe; SESSDATA=686ee018%2C1709376822%2C285fa%2A92t8Gw_WX6Qn28x-kxXiMZ_MuyrQaE71j1fMmLruLwXO9kFnz0oJSzUYw4FvvNayLgXTp25AAAXgA; bili_jct=3d9d106c9dccf05ff06826424515af4e; DedeUserID=516093694; DedeUserID__ckMd5=7d65b431d2712517; CURRENT_QUALITY=80; bg_view_28884=770277; bp_video_offset_516093694=838133520311255063; b_lsid=B1953F3E_18A78D02DB7; sid=6jhdk4r3; PVID=3',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36'
    }
    response=requests.get(url=url,headers=headers)
    response.encoding='utf-8'
    #print(response.text)
    text_list=re.findall('bvid:"(.*?)"',response.text)
    for index in text_list:
        # 爬取视频弹幕的url
        vidio_url='https://www.ibilibili.com/video/'+index
        #print(vidio_url)
        response_1=requests.get(url=vidio_url)
        response_1.encoding='utf-8'
        #print(response_1.text)
        danmu_url=re.findall('<a href="(.*?)"  class="btn btn-default" target="_blank">弹幕</a>',response_1.text)
        print(danmu_url)
        response_2=requests.get(url=danmu_url[0])
        response_2.encoding='utf-8'
        danmu_list=re.findall('</d><d p=".*?">(.*?)</d>',response_2.text)
        for i in danmu_list:
            with open('弹幕.text', mode='a', encoding='utf-8') as f:
                f.write(i)
                f.write('\n')
                print(i)
                if i in cnts:
                    cnts[i]+=1
                else:
                    cnts[i]=1
dic = {}
dic.update({k: cnts[k] for k in heapq.nlargest(20, cnts, key=cnts.get)})
arr=[]
for key,value in dic.items():
    arr.append((key,value))
def write_lines_excel(arr):#将排名前20的弹幕写入表格
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('new')
    sheet.cell(1, 1, '数量排名前20的弹幕内容')
    sheet.cell(1, 2, '弹幕出现次数')
    for index, row in enumerate(arr):
        for col in range(len(row)):
            sheet.cell(index + 2, col + 1, row[col])
    work_book.save('前20弹幕.xlsx')
write_lines_excel(arr)
print(dic)
                #print(i)
        #print(danmu_url)