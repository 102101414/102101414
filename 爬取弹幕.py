import requests
import re
import heapq
import openpyxl
cnts={}
def get_vidio_url():
    vidio_url=[]
    for page in range(1,11):
        # 获取一页视频网页的url
        if page==1:
            url=f'https://search.bilibili.com/all?keyword=日本核污染水排海'
        else:
            url=f'https://search.bilibili.com/all?vt=68973445&keyword=日本核污染水排海&page={page}'
        headers = {
        'cookie':'nostalgia_conf=-1; _uuid=DBA108CF10-7B18-31D2-F235-6A7D287698B813584infoc; buvid3=E13C5F69-3CDE-3643-28D3-53551C09952A14835infoc; b_nut=1666596615; buvid4=EF91F08C-3200-880E-183C-32A8153B8D3314835-022102415-V35mpzdvTWSkhBz1UroM1g%3D%3D; i-wanna-go-back=-1; b_ut=7; is-2022-channel=1; FEED_LIVE_VERSION=V8; header_theme_version=CLOSE; home_feed_column=5; CURRENT_BLACKGAP=0; rpdid=|(u))kkYuuu|0J\'uY))J~kuum; CURRENT_FNVAL=4048; browser_resolution=1494-789; fingerprint=722658bd40bab4c1be58d7e8b4e4fcfe; buvid_fp_plain=undefined; buvid_fp=722658bd40bab4c1be58d7e8b4e4fcfe; SESSDATA=686ee018%2C1709376822%2C285fa%2A92t8Gw_WX6Qn28x-kxXiMZ_MuyrQaE71j1fMmLruLwXO9kFnz0oJSzUYw4FvvNayLgXTp25AAAXgA; bili_jct=3d9d106c9dccf05ff06826424515af4e; DedeUserID=516093694; DedeUserID__ckMd5=7d65b431d2712517; CURRENT_QUALITY=80; bg_view_28884=770277; bp_video_offset_516093694=838133520311255063; b_lsid=B1953F3E_18A78D02DB7; sid=6jhdk4r3; PVID=3',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36'
        }
        response=requests.get(url=url,headers=headers)
        response.encoding = 'utf-8'
        # 获取一页下面每个视屏的url
        text_list=re.findall('bvid:"(.*?)"',response.text)
        vidio_url+=text_list
    return vidio_url

def get_interface_url(vidio_url):
    interface_url=[]
    for index in vidio_url:
        # 通过视频的url得到视频弹幕地址的url
        url='https://www.ibilibili.com/video/'+index
        interface_url.append(url)

    return interface_url
def get_barrage_url(interface_url):
    barrage_url=[]
    for url in interface_url:
        response = requests.get(url=url)
        response.encoding = 'utf-8'
        new_url = re.findall('<a href="(.*?)"  class="btn btn-default" target="_blank">弹幕</a>', response.text)
        barrage_url +=new_url
    return barrage_url

def get_barrage(barrage_url):
    for url in barrage_url:
        response = requests.get(url=url)
        response.encoding = 'utf-8'
        barrage_list = re.findall('</d><d p=".*?">(.*?)</d>', response.text)
        for index in barrage_list:
            with open('弹幕.text', mode='a', encoding='utf-8') as f:
                f.write(index)
                f.write('\n')
                #获取每种弹幕的数量
                if index in cnts:
                    cnts[index] += 1
                else:
                    cnts[index] = 1

def write_lines_excel(arr):
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('arrage')
    sheet.cell(1, 1, '数量排名前20的弹幕内容')
    sheet.cell(1, 2, '弹幕出现次数')
    for index, row in enumerate(arr):
        for col in range(len(row)):
            sheet.cell(index + 2, col + 1, row[col])
    work_book.save('前20弹幕.xlsx')
#获取300个视频的url,并存在vidio_url列表中
vidio_url=get_vidio_url()
#获取弹幕地址所在html的url,并存在interface_url列表中
interface_url=get_interface_url(vidio_url)
#获取弹幕url，并存在barrage_url中
barrage_url=get_barrage_url(interface_url)
#获取每一个弹幕url下的弹幕，并保存在‘弹幕.text’zhong
get_barrage(barrage_url)
#将得到的弹幕按照数量多少进行排序
dic = {}
dic.update({k: cnts[k] for k in heapq.nlargest(20, cnts, key=cnts.get)})
arr=[]
for key,value in dic.items():
    arr.append((key,value))
#将前20的弹幕生成表格
write_lines_excel(arr)
print(dic)